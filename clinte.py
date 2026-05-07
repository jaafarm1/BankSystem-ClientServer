# ------------------------------------------------------------
# client.py - Bank Client with Professional Menus & Notifications
# ------------------------------------------------------------
import socket
import json
import getpass
import sys
import os
from datetime import datetime

# Import multi-language support
from languages import t, set_language, get_language, LANGUAGES

# Try to import libraries with error handling
try:
    from plyer import notification
    NOTIFICATIONS_AVAILABLE = True
except ImportError:
    NOTIFICATIONS_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ------------------------------------------------------------
# Helper function to get value from dict or tuple/list
# ------------------------------------------------------------
def get_value(data, dict_key=None, tuple_index=0, default=''):
    """Extract value from data whether it's dict or tuple/list"""
    if data is None:
        return default
    elif isinstance(data, dict):
        if dict_key and dict_key in data:
            return data[dict_key]
        return default
    elif isinstance(data, (list, tuple)):
        if len(data) > tuple_index:
            return data[tuple_index]
        return default
    return data

# ------------------------------------------------------------
# Notification Function
# ------------------------------------------------------------
def send_notification(title, message):
    """Send desktop notification with proper language"""
    if NOTIFICATIONS_AVAILABLE:
        try:
            current_lang = get_language()
            
            if current_lang == 'ar':
                notif_titles = {
                    'login': 'تسجيل دخول',
                    'logout': 'تسجيل خروج',
                    'balance': 'الرصيد الحالي',
                    'deposit': 'إيداع ناجح',
                    'withdraw': 'سحب ناجح',
                    'transfer': 'تحويل ناجح',
                    'change_pass': 'تغيير كلمة المرور',
                    'backup': 'نسخة احتياطية',
                    'export_excel': 'تصدير Excel',
                    'create_account': 'حساب جديد',
                    'create_user': 'مستخدم جديد',
                    'delete_account': 'حذف حساب',
                    'delete_user': 'حذف مستخدم',
                    'set_limit': 'حد السحب اليومي',
                    'update_rates': 'تحديث الأسعار',
                    'graph_report': 'تقرير رسوم بيانية'
                }
                notif_title = notif_titles.get(title, title)
            else:
                notif_titles = {
                    'login': 'Login',
                    'logout': 'Logout',
                    'balance': 'Current Balance',
                    'deposit': 'Deposit Success',
                    'withdraw': 'Withdrawal Success',
                    'transfer': 'Transfer Success',
                    'change_pass': 'Password Changed',
                    'backup': 'Backup',
                    'export_excel': 'Excel Export',
                    'create_account': 'New Account',
                    'create_user': 'New User',
                    'delete_account': 'Account Deleted',
                    'delete_user': 'User Deleted',
                    'set_limit': 'Daily Limit',
                    'update_rates': 'Rates Updated',
                    'graph_report': 'Graph Report'
                }
                notif_title = notif_titles.get(title, title)
            
            notification.notify(
                title=notif_title,
                message=message,
                app_name="Bank System",
                timeout=3
            )
        except Exception as e:
            pass

# ------------------------------------------------------------
# Bank Client Class
# ------------------------------------------------------------
class BankClient:
    def __init__(self, host='127.0.0.1', port=5555):
        self.host = host
        self.port = port
        self.socket = None
        self.authenticated = False
        self.username = None
        self.role = None
    
    def connect(self):
        """Connect to server"""
        try:
            self.socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.socket.connect((self.host, self.port))
            return True
        except Exception as e:
            print(f"[ERROR] {t('connecting_error')}: {e}")
            return False
    
    def send_request(self, request):
        """Send request to server and get response"""
        try:
            self.socket.send(json.dumps(request, ensure_ascii=False).encode('utf-8'))
            response = json.loads(self.socket.recv(4096).decode('utf-8'))
            return response
        except Exception as e:
            print(f"[ERROR] {t('connecting_error')}: {e}")
            return None
    
    def login(self, username, password):
        """Login to system"""
        request = {
            'action': 'login',
            'username': username,
            'password': password
        }
        response = self.send_request(request)
        if response and response.get('status') == 'success':
            self.authenticated = True
            self.username = username
            self.role = response.get('role')
            return True, response.get('message')
        return False, response.get('message') if response else t('connecting_error')
    
    def logout(self):
        """Logout from system"""
        if self.authenticated:
            try:
                request = {'action': 'logout'}
                self.send_request(request)
            except:
                pass
        self.authenticated = False
        self.username = None
        self.role = None
    
    def get_balance(self, account_number):
        request = {'action': 'balance', 'account_number': account_number}
        return self.send_request(request)
    
    def deposit(self, account_number, amount):
        request = {'action': 'deposit', 'account_number': account_number, 'amount': amount}
        return self.send_request(request)
    
    def withdraw(self, account_number, amount):
        request = {'action': 'withdraw', 'account_number': account_number, 'amount': amount}
        return self.send_request(request)
    
    def transfer(self, from_account, to_account, amount):
        request = {'action': 'transfer', 'from_account': from_account, 'to_account': to_account, 'amount': amount}
        return self.send_request(request)
    
    def get_history(self, account_number):
        request = {'action': 'history', 'account_number': account_number}
        return self.send_request(request)
    
    def create_user(self, username, password, role):
        request = {'action': 'create_user', 'username': username, 'password': password, 'role': role}
        return self.send_request(request)
    
    def create_account(self, username, account_number, initial_balance):
        request = {'action': 'create_account', 'username': username, 'account_number': account_number, 'initial_balance': initial_balance}
        return self.send_request(request)
    
    def list_users(self):
        request = {'action': 'list_users'}
        return self.send_request(request)
    
    def admin_report(self):
        request = {'action': 'admin_report'}
        return self.send_request(request)
    
    def graph_report(self):
        request = {'action': 'graph_report'}
        return self.send_request(request)
    
    def delete_account(self, account_number):
        request = {'action': 'delete_account', 'account_number': account_number}
        return self.send_request(request)
    
    def delete_user(self, username):
        request = {'action': 'delete_user', 'username': username}
        return self.send_request(request)
    
    def backup(self):
        request = {'action': 'backup'}
        return self.send_request(request)
    
    def change_password(self, username, old_password, new_password):
        request = {'action': 'change_password', 'username': username, 'old_password': old_password, 'new_password': new_password}
        return self.send_request(request)
    
    def set_daily_limit(self, account_number, limit):
        request = {'action': 'set_daily_limit', 'account_number': account_number, 'limit': limit}
        return self.send_request(request)
    
    def get_daily_info(self, account_number):
        request = {'action': 'get_daily_info', 'account_number': account_number}
        return self.send_request(request)
    
    def convert_currency(self, amount, from_currency, to_currency):
        request = {'action': 'convert_currency', 'amount': amount, 'from_currency': from_currency, 'to_currency': to_currency}
        return self.send_request(request)
    
    def balance_multiple_currencies(self, account_number):
        request = {'action': 'balance_multiple_currencies', 'account_number': account_number}
        return self.send_request(request)
    
    def get_exchange_rates(self):
        request = {'action': 'get_exchange_rates'}
        return self.send_request(request)
    
    def update_exchange_rates(self, rates):
        request = {'action': 'update_exchange_rates', 'rates': rates}
        return self.send_request(request)
    
    def export_to_excel_file(self, filename=None):
        """Export data to Excel file"""
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl library not installed"
        
        if filename is None:
            filename = f"bank_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(t('creating_excel'))
        
        response = self.admin_report()
        if not response or response.get('status') != 'success':
            return False, response.get('message') if response else 'Error'
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Accounts
        ws1 = wb.active
        ws1.title = "Accounts"
        headers1 = ['Account Number', 'Username', 'Role', 'Balance (JOD)', 'Created Date']
        ws1.append(headers1)
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for acc in response.get('accounts', []):
            # Support both dict and tuple formats
            if isinstance(acc, dict):
                ws1.append([
                    acc.get('account_number', ''),
                    acc.get('username', ''),
                    acc.get('role', ''),
                    acc.get('balance', 0),
                    acc.get('created_date', '')
                ])
            else:
                ws1.append([
                    acc[0] if len(acc) > 0 else '',
                    acc[1] if len(acc) > 1 else '',
                    acc[2] if len(acc) > 2 else '',
                    acc[3] if len(acc) > 3 else 0,
                    acc[5] if len(acc) > 5 else ''
                ])
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws1.column_dimensions[col].width = 18
        
        # Sheet 2: Statistics
        ws2 = wb.create_sheet("Statistics")
        ws2.append(['Indicator', 'Value'])
        ws2.append(['Total Users', response.get('users_count', 0)])
        ws2.append(['Total Accounts', response.get('accounts_count', 0)])
        ws2.append(['Total Balance (JOD)', response.get('total_balance', 0)])
        ws2.append(['', ''])
        ws2.append(['Transaction Type', 'Count', 'Total Amount'])
        for stat in response.get('stats', []):
            if isinstance(stat, dict):
                type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(stat.get('type', ''), stat.get('type', ''))
                ws2.append([type_en, stat.get('count', 0), stat.get('total', 0)])
            else:
                type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(stat[0] if len(stat) > 0 else '', stat[0] if len(stat) > 0 else '')
                ws2.append([type_en, stat[1] if len(stat) > 1 else 0, stat[2] if len(stat) > 2 else 0])
        
        for col in ['A', 'B', 'C']:
            ws2.column_dimensions[col].width = 20
        
        # Sheet 3: Most Active Accounts
        ws3 = wb.create_sheet("Most Active Accounts")
        ws3.append(['Account Number', 'Transaction Count'])
        for acc in response.get('active_accounts', []):
            if isinstance(acc, dict):
                ws3.append([acc.get('account_number', ''), acc.get('transaction_count', 0)])
            else:
                ws3.append([acc[0] if len(acc) > 0 else '', acc[1] if len(acc) > 1 else 0])
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 20
        
        # Sheet 4: Daily Summary
        ws4 = wb.create_sheet("Daily Summary")
        ws4.append(['Transaction Type', 'Total Amount'])
        for d in response.get('daily_summary', []):
            if isinstance(d, dict):
                type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(d.get('type', ''), d.get('type', ''))
                ws4.append([type_en, d.get('total', 0)])
            else:
                type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(d[0] if len(d) > 0 else '', d[0] if len(d) > 0 else '')
                ws4.append([type_en, d[1] if len(d) > 1 else 0])
        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 20
        
        wb.save(filename)
        return True, filename

# ------------------------------------------------------------
# Graph Functions
# ------------------------------------------------------------
def show_graph_report(data):
    """Display graphical report for admin"""
    
    if not MATPLOTLIB_AVAILABLE:
        print("\n[x] matplotlib library not installed")
        return
    
    if not data:
        print("\n[x] No data available")
        return
    
    accounts = data.get('accounts', [])
    users = data.get('users', [])
    stats = data.get('stats', [])
    active_accounts = data.get('active_accounts', [])
    
    if not accounts:
        print("\n[x] No accounts to display")
        return
    
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'Tahoma']
    plt.rcParams['axes.unicode_minus'] = False
    
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle('Bank Report - Charts', fontsize=16, fontweight='bold')
    
    # 1. Account Balances
    # Support both dict and tuple formats
    acc_names = []
    balances = []
    for acc in accounts[:8]:
        if isinstance(acc, dict):
            acc_names.append(str(acc.get('account_number', '')))
            balances.append(acc.get('balance', 0))
        else:
            acc_names.append(str(acc[0] if len(acc) > 0 else ''))
            balances.append(acc[3] if len(acc) > 3 else 0)
    
    bars = axes[0, 0].bar(acc_names, balances, color='steelblue', edgecolor='navy')
    axes[0, 0].set_title('Account Balances (JOD)', fontsize=12, pad=10)
    axes[0, 0].set_xlabel('Account Number', fontsize=10)
    axes[0, 0].set_ylabel('Balance (JOD)', fontsize=10)
    axes[0, 0].tick_params(axis='x', rotation=45, labelsize=9)
    
    for bar, bal in zip(bars, balances):
        axes[0, 0].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 100,
                       f'{bal:,.0f}', ha='center', va='bottom', fontsize=8)
    
    # 2. User Distribution
    if users:
        roles = {}
        for u in users:
            if isinstance(u, dict):
                role = u.get('role', '')
            else:
                role = u[2] if len(u) > 2 else ''
            roles[role] = roles.get(role, 0) + 1
        
        labels = {'admin': 'Admin', 'teller': 'Teller', 'customer': 'Customer'}
        role_labels = [labels.get(k, k) for k in roles.keys()]
        sizes = list(roles.values())
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1']
        
        axes[0, 1].pie(sizes, labels=role_labels, autopct='%1.1f%%', 
                       colors=colors[:len(sizes)], startangle=90)
        axes[0, 1].set_title('User Distribution by Role', fontsize=12, pad=10)
    
    # 3. Transaction Statistics
    if stats:
        stat_types = []
        stat_totals = []
        for stat in stats:
            if isinstance(stat, dict):
                type_val = stat.get('type', '')
                type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(type_val, type_val)
                stat_types.append(type_en)
                stat_totals.append(stat.get('total', 0))
            else:
                type_val = stat[0] if len(stat) > 0 else ''
                type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(type_val, type_val)
                stat_types.append(type_en)
                stat_totals.append(stat[2] if len(stat) > 2 and stat[2] else 0)
        
        colors_bar = ['green' if 'Deposit' in t else 'red' if 'Withdraw' in t else 'orange' for t in stat_types]
        axes[1, 0].barh(stat_types, stat_totals, color=colors_bar)
        axes[1, 0].set_title('Total Transactions by Type', fontsize=12, pad=10)
        axes[1, 0].set_xlabel('Total Amount (JOD)', fontsize=10)
        
        for bar, total in zip(axes[1, 0].patches, stat_totals):
            axes[1, 0].text(bar.get_width() + 50, bar.get_y() + bar.get_height()/2,
                           f'{total:,.0f}', ha='left', va='center', fontsize=8)
    
    # 4. Most Active Accounts
    if active_accounts:
        active_accs = []
        active_counts = []
        for a in active_accounts[:5]:
            if isinstance(a, dict):
                active_accs.append(str(a.get('account_number', '')))
                active_counts.append(a.get('transaction_count', 0))
            else:
                active_accs.append(str(a[0] if len(a) > 0 else ''))
                active_counts.append(a[1] if len(a) > 1 else 0)
        
        axes[1, 1].bar(active_accs, active_counts, color='purple', alpha=0.7)
        axes[1, 1].set_title('Most Active Accounts', fontsize=12, pad=10)
        axes[1, 1].set_xlabel('Account Number', fontsize=10)
        axes[1, 1].set_ylabel('Number of Transactions', fontsize=10)
        axes[1, 1].tick_params(axis='x', rotation=45)
        
        for bar, count in zip(axes[1, 1].patches, active_counts):
            axes[1, 1].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                           f'{count}', ha='center', va='bottom', fontsize=8)
    
    plt.tight_layout()
    plt.show()

# ------------------------------------------------------------
# Language Selection
# ------------------------------------------------------------
def select_language():
    """Show language selection menu"""
    clear_screen()
    print("=" * 55)
    print("   Select Language / اختر اللغة")
    print("=" * 55)
    print(" 1. English")
    print(" 2. العربية")
    print("-" * 55)
    
    choice = input("Choose / اختر (1 or 2): ")
    
    if choice == '1':
        set_language('en')
    elif choice == '2':
        set_language('ar')
    else:
        set_language('en')
    
    clear_screen()

# ------------------------------------------------------------
# User Interface Menus
# ------------------------------------------------------------
def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def print_header(title):
    print("=" * 55)
    print(f"  {title}")
    print("=" * 55)

def customer_menu(client, account_number):
    """Customer Menu"""
    while True:
        clear_screen()
        print_header(f"{t('welcome')} {client.username} ({t('customer_role')})")
        print(f" 1. {t('check_balance')}")
        print(f" 2. {t('deposit')}")
        print(f" 3. {t('withdraw')}")
        print(f" 4. {t('transfer')}")
        print(f" 5. {t('history')}")
        print(f" 6. {t('multi_currency')}")
        print(f" 7. {t('convert')}")
        print(f" 8. {t('daily_limit')}")
        print(f" 9. {t('change_pass')}")
        print(f"10. {t('logout_btn')}")
        print("=" * 55)
        
        choice = input("🔹 Choose / اختر العملية (1-10): ")
        
        if choice == '1':
            response = client.get_balance(account_number)
            if response and response.get('status') == 'success':
                msg = f"{response['balance']:,.2f} JOD"
                print(f"\n✅ {t('balance')} {msg}")
                send_notification('balance', msg)
            else:
                print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '2':
            try:
                amount = float(input(f"{t('amount')} "))
                response = client.deposit(account_number, amount)
                if response and response.get('status') == 'success':
                    msg = f"{amount:,.2f} JOD"
                    print(f"\n✅ {response['message']}")
                    print(f"    {t('balance')} {response['balance']:,.2f} {t('jod')}")
                    send_notification('deposit', msg)
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '3':
            try:
                amount = float(input(f"{t('amount')} "))
                response = client.withdraw(account_number, amount)
                if response and response.get('status') == 'success':
                    msg = f"{amount:,.2f} JOD"
                    print(f"\n✅ {response['message']}")
                    print(f"    {t('balance')} {response['balance']:,.2f} {t('jod')}")
                    send_notification('withdraw', msg)
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '4':
            to_account = input(f"{t('to_account')} ")
            try:
                amount = float(input(f"{t('amount')} "))
                response = client.transfer(account_number, to_account, amount)
                if response and response.get('status') == 'success':
                    msg = f"{amount:,.2f} JOD to {to_account}"
                    print(f"\n✅ {response['message']}")
                    print(f"    {t('balance')} {response['balance']:,.2f} {t('jod')}")
                    send_notification('transfer', msg)
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '5':
            response = client.get_history(account_number)
            if response and response.get('status') == 'success':
                history = response.get('history', [])
                if history:
                    print(f"\n{t('history_title')}")
                    print("-" * 65)
                    print(f"   {t('type'):<12} {t('amount'):<12} {t('target'):<15} {t('date'):<20}")
                    print("-" * 65)
                    for trans in history:
                        # Support both dict and tuple
                        if isinstance(trans, dict):
                            trans_type = trans.get('transaction_type', '')
                            type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(trans_type, trans_type)
                            amount_val = trans.get('amount', 0)
                            target = trans.get('to_account', '-')
                            date_val = trans.get('transaction_date', '')
                        else:
                            type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(trans[1] if len(trans) > 1 else '', trans[1] if len(trans) > 1 else '')
                            amount_val = trans[2] if len(trans) > 2 else 0
                            target = trans[3] if len(trans) > 3 else '-'
                            date_val = trans[4] if len(trans) > 4 else ''
                        print(f"   {type_en:<12} {amount_val:<12,} {target:<15} {date_val:<20}")
                else:
                    print("\n📭 No transactions found")
            else:
                print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '6':
            response = client.balance_multiple_currencies(account_number)
            if response and response.get('status') == 'success':
                print(f"\n{t('balance_jod')} {response['balance_jod']:,.2f} JD")
                print(f"\n{t('other_currencies')}")
                for currency, amount in response.get('balances', {}).items():
                    print(f"   {currency}: {amount:,.2f}")
            else:
                print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '7':
            try:
                amount = float(input(f"{t('amount')} "))
                print(f"\n{t('available_currencies')}")
                from_curr = input(f"{t('from_currency')} ").upper()
                to_curr = input(f"{t('to_currency')} ").upper()
                response = client.convert_currency(amount, from_curr, to_curr)
                if response and response.get('status') == 'success':
                    print(f"\n✅ {t('convert_result')} {amount:,.2f} {from_curr} = {response['converted_amount']:,.2f} {to_curr}")
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '8':
            response = client.get_daily_info(account_number)
            if response and response.get('status') == 'success':
                print(f"\n{t('daily_limit_text')} {response['limit']:,.2f} {t('jod')}")
                print(f"{t('withdrawn_today')} {response['withdrawn']:,.2f} {t('jod')}")
                print(f"{t('remaining')} {response['remaining']:,.2f} {t('jod')}")
            else:
                print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '9':
            old_pass = getpass.getpass(f"{t('old_password')} ")
            new_pass = getpass.getpass(f"{t('new_password_text')} ")
            confirm = getpass.getpass(f"{t('confirm_new')} ")
            if new_pass != confirm:
                print(f"\n❌ {t('password_mismatch')}")
            else:
                response = client.change_password(client.username, old_pass, new_pass)
                if response and response.get('status') == 'success':
                    print(f"\n✅ {response['message']}")
                    send_notification('change_pass', 'Password changed successfully')
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '10':
            print(f"\n{t('logout')}")
            client.logout()
            print(f"✅ {t('logged_out')}")
            send_notification('logout', 'Logged out successfully')
            input(f"\n{t('press_enter')}")
            return
        
        else:
            print(f"\n❌ {t('invalid_choice')}")
        
        input(f"\n{t('press_enter')}")

def teller_menu(client):
    """Teller Menu"""
    while True:
        clear_screen()
        print_header(f"{t('welcome')} {client.username} ({t('teller_role')})")
        print(f" 1. {t('check_balance_teller')}")
        print(f" 2. {t('deposit_teller')}")
        print(f" 3. {t('withdraw_teller')}")
        print(f" 4. {t('transfer_teller')}")
        print(f" 5. {t('history_teller')}")
        print(f" 6. {t('open_account')}")
        print(f" 7. {t('create_user')}")
        print(f" 8. {t('multi_currency_teller')}")
        print(f" 9. {t('convert_teller')}")
        print(f"10. {t('change_pass_teller')}")
        print(f"11. {t('logout_teller')}")
        print("=" * 55)
        
        choice = input("🔹 Choose / اختر العملية (1-11): ")
        
        if choice == '1':
            acc = input(f"{t('account_number')} ")
            response = client.get_balance(acc)
            if response and response.get('status') == 'success':
                print(f"\n✅ {t('balance')} {response['balance']:,.2f} {t('jod')}")
            else:
                print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '2':
            acc = input(f"{t('account_number')} ")
            try:
                amount = float(input(f"{t('amount')} "))
                response = client.deposit(acc, amount)
                if response and response.get('status') == 'success':
                    print(f"\n✅ {response['message']}")
                    send_notification('deposit', f"{amount:,.2f} JOD to {acc}")
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '3':
            acc = input(f"{t('account_number')} ")
            try:
                amount = float(input(f"{t('amount')} "))
                response = client.withdraw(acc, amount)
                if response and response.get('status') == 'success':
                    print(f"\n✅ {response['message']}")
                    send_notification('withdraw', f"{amount:,.2f} JOD from {acc}")
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '4':
            from_acc = input(f"{t('from_account')} ")
            to_acc = input(f"{t('to_account')} ")
            try:
                amount = float(input(f"{t('amount')} "))
                response = client.transfer(from_acc, to_acc, amount)
                if response and response.get('status') == 'success':
                    print(f"\n✅ {response['message']}")
                    send_notification('transfer', f"{amount:,.2f} JOD from {from_acc} to {to_acc}")
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '5':
            acc = input(f"{t('account_number')} ")
            response = client.get_history(acc)
            if response and response.get('status') == 'success':
                history = response.get('history', [])
                if history:
                    print(f"\n{t('history_title')}")
                    for trans in history:
                        if isinstance(trans, dict):
                            type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(trans.get('transaction_type', ''), trans.get('transaction_type', ''))
                            amount_val = trans.get('amount', 0)
                            date_val = trans.get('transaction_date', '')
                        else:
                            type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(trans[1] if len(trans) > 1 else '', trans[1] if len(trans) > 1 else '')
                            amount_val = trans[2] if len(trans) > 2 else 0
                            date_val = trans[4] if len(trans) > 4 else ''
                        print(f"   {type_en} | {amount_val} | {date_val}")
                else:
                    print("\n📭 No transactions found")
            else:
                print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '6':
            username = input(f"{t('new_username')} ")
            acc_num = input(f"{t('new_account')} ")
            try:
                balance = float(input(f"{t('initial_balance')} "))
                response = client.create_account(username, acc_num, balance)
                if response and response.get('status') == 'success':
                    print(f"\n✅ {response['message']}")
                    send_notification('create_account', f"{acc_num} for {username}")
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '7':
            new_user = input(f"{t('new_username')} ")
            new_pass = getpass.getpass(f"{t('new_password')} ")
            print(f"\n{t('role_type')} customer, teller")
            role = input(f"{t('role_type')} ")
            if role not in ['customer', 'teller']:
                print(f"\n❌ {t('invalid_choice')}")
            else:
                response = client.create_user(new_user, new_pass, role)
                if response and response.get('status') == 'success':
                    print(f"\n✅ {response['message']}")
                    send_notification('create_user', f"{new_user} as {role}")
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '8':
            acc = input(f"{t('account_number')} ")
            response = client.balance_multiple_currencies(acc)
            if response and response.get('status') == 'success':
                print(f"\n{t('balance_jod')} {response['balance_jod']:,.2f} JD")
                print(f"\n{t('other_currencies')}")
                for curr, amt in response.get('balances', {}).items():
                    print(f"   {curr}: {amt:,.2f}")
            else:
                print(f"\n❌ {response.get('message') if response else t('error')}")
        
        elif choice == '9':
            try:
                amount = float(input(f"{t('amount')} "))
                from_c = input(f"{t('from_currency')} ").upper()
                to_c = input(f"{t('to_currency')} ").upper()
                response = client.convert_currency(amount, from_c, to_c)
                if response and response.get('status') == 'success':
                    print(f"\n✅ {t('convert_result')} {amount:,.2f} {from_c} = {response['converted_amount']:,.2f} {to_c}")
                else:
                    print(f"\n❌ {response.get('message') if response else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '10':
            old = getpass.getpass(f"{t('old_password')} ")
            new = getpass.getpass(f"{t('new_password_text')} ")
            conf = getpass.getpass(f"{t('confirm_new')} ")
            if new != conf:
                print(f"\n❌ {t('password_mismatch')}")
            else:
                resp = client.change_password(client.username, old, new)
                if resp and resp.get('status') == 'success':
                    print(f"\n✅ {resp['message']}")
                    send_notification('change_pass', 'Password changed successfully')
                else:
                    print(f"\n❌ {resp.get('message') if resp else t('error')}")
        
        elif choice == '11':
            print(f"\n{t('logout')}")
            client.logout()
            print(f"✅ {t('logged_out')}")
            send_notification('logout', 'Logged out successfully')
            input(f"\n{t('press_enter')}")
            return
        
        else:
            print(f"\n❌ {t('invalid_choice')}")
        
        input(f"\n{t('press_enter')}")

def admin_menu(client):
    """Admin Menu"""
    while True:
        clear_screen()
        print_header(f"{t('welcome')} {client.username} ({t('admin_role')})")
        print(f" 1. {t('account_mgmt')}")
        print(f" 2. {t('user_mgmt')}")
        print(f" 3. {t('full_report')}")
        print(f" 4. {t('graph_report')}")
        print(f" 5. {t('backup')}")
        print(f" 6. {t('set_limit')}")
        print(f" 7. {t('update_rates')}")
        print(f" 8. {t('export_excel')}")
        print(f" 9. {t('currencies')}")
        print(f"10. {t('change_pass_admin')}")
        print(f"11. {t('logout_admin')}")
        print("=" * 55)
        
        choice = input("🔹 Choose / اختر العملية (1-11): ")
        
        if choice == '1':
            print(f"\n 1. {t('view_accounts')}")
            print(f" 2. {t('create_new_account')}")
            print(f" 3. {t('delete_account')}")
            print("-" * 40)
            sub = input("Choose / اختر (1-3): ")
            
            if sub == '1':
                resp = client.admin_report()
                if resp and resp.get('status') == 'success':
                    accounts = resp.get('accounts', [])
                    if accounts:
                        print(f"\n📋 {t('account_list')}")
                        print("-" * 55)
                        print(f"   {'Account Number':<15} {'Username':<15} {'Balance (JOD)':>15}")
                        print("-" * 55)
                        for acc in accounts:
                            # Support both dict and tuple formats
                            if isinstance(acc, dict):
                                acc_num = acc.get('account_number', '')
                                username = acc.get('username', '')
                                balance = acc.get('balance', 0)
                            else:
                                acc_num = acc[0] if len(acc) > 0 else ''
                                username = acc[1] if len(acc) > 1 else ''
                                balance = acc[3] if len(acc) > 3 else 0
                            print(f"   {acc_num:<15} {username:<15} {balance:>15,.2f}")
                        print("-" * 55)
                        print(f"   Total Accounts: {len(accounts)}")
                    else:
                        print(f"\n📭 {t('no_accounts')}")
                else:
                    print(f"\n❌ {resp.get('message') if resp else t('error')}")
            
            elif sub == '2':
                print(f"\n📝 {t('user_must_exist')}")
                print("-" * 50)
                
                users_resp = client.list_users()
                if users_resp and users_resp.get('status') == 'success':
                    users = users_resp.get('users', [])
                    if users:
                        print(f"\n📋 {t('available_users')}")
                        print("-" * 40)
                        for u in users:
                            if isinstance(u, dict):
                                username = u.get('username', '')
                                role = u.get('role', '')
                            else:
                                username = u[1] if len(u) > 1 else ''
                                role = u[2] if len(u) > 2 else ''
                            print(f"   • {username} (Role: {role})")
                        print("-" * 40)
                print("-" * 50)
                
                username = input(f"\n{t('enter_username')} ")
                acc_num = input(f"{t('new_account')} (e.g., ACC002): ")
                
                try:
                    balance = float(input(f"{t('initial_balance')} (JOD): "))
                    if balance < 0:
                        print(f"\n❌ {t('invalid_amount')}")
                    else:
                        resp = client.create_account(username, acc_num, balance)
                        if resp and resp.get('status') == 'success':
                            print(f"\n✅ {resp['message']}")
                            print(f"   📌 Details: Account {acc_num} | Balance: {balance:.2f} JOD")
                            send_notification('create_account', f"{acc_num} for {username}")
                        else:
                            print(f"\n❌ {resp.get('message') if resp else t('error')}")
                except ValueError:
                    print(f"\n❌ {t('invalid_amount')}")
            
            elif sub == '3':
                acc = input(f"{t('account_number')} ")
                confirm = input(f"Are you sure you want to delete account {acc}? (y/n): ")
                if confirm.lower() == 'y':
                    resp = client.delete_account(acc)
                    if resp and resp.get('status') == 'success':
                        print(f"\n✅ {resp['message']}")
                        send_notification('delete_account', f"{acc}")
                    else:
                        print(f"\n❌ {resp.get('message') if resp else t('error')}")
                else:
                    print("\n❌ Operation cancelled")
            
            else:
                print(f"\n❌ {t('invalid_choice')}")
        
        elif choice == '2':
            print(f"\n 1. {t('view_users')}")
            print(f" 2. {t('delete_user')}")
            print(f" 3. {t('create_new_user')}")
            print("-" * 40)
            sub = input("Choose / اختر (1-3): ")
            
            if sub == '1':
                resp = client.list_users()
                if resp and resp.get('status') == 'success':
                    users = resp.get('users', [])
                    if users:
                        print(f"\n📋 {t('user_list')}")
                        print("-" * 65)
                        print(f"   {'ID':<5} {'Username':<15} {'Role':<12} {'Created Date':<20}")
                        print("-" * 65)
                        for u in users:
                            if isinstance(u, dict):
                                uid = u.get('user_id', '')
                                username = u.get('username', '')
                                role = u.get('role', '')
                                created = u.get('created_date', '')
                            else:
                                uid = u[0] if len(u) > 0 else ''
                                username = u[1] if len(u) > 1 else ''
                                role = u[2] if len(u) > 2 else ''
                                created = u[3] if len(u) > 3 else ''
                            print(f"   {uid:<5} {username:<15} {role:<12} {created:<20}")
                        print("-" * 65)
                        print(f"   Total Users: {len(users)}")
                    else:
                        print(f"\n📭 {t('no_users')}")
                else:
                    print(f"\n❌ {resp.get('message') if resp else t('error')}")
            
            elif sub == '2':
                username = input(f"{t('enter_username')} ")
                if username == 'admin':
                    print(f"\n❌ {t('cannot_delete_admin')}")
                else:
                    confirm = input(f"Are you sure you want to delete user {username}? (y/n): ")
                    if confirm.lower() == 'y':
                        resp = client.delete_user(username)
                        if resp and resp.get('status') == 'success':
                            print(f"\n✅ {resp['message']}")
                            send_notification('delete_user', f"{username}")
                        else:
                            print(f"\n❌ {resp.get('message') if resp else t('error')}")
                    else:
                        print("\n❌ Operation cancelled")
            
            elif sub == '3':
                new_user = input(f"{t('new_username')} ")
                new_pass = getpass.getpass(f"{t('new_password')} ")
                print(f"\n{t('role_type')}: customer, teller, admin")
                role = input(f"{t('role_type')} ")
                if role not in ['customer', 'teller', 'admin']:
                    print(f"\n❌ {t('invalid_choice')}")
                else:
                    resp = client.create_user(new_user, new_pass, role)
                    if resp and resp.get('status') == 'success':
                        print(f"\n✅ {resp['message']}")
                        print(f"   📌 User: {new_user} | Role: {role}")
                        send_notification('create_user', f"{new_user} as {role}")
                    else:
                        print(f"\n❌ {resp.get('message') if resp else t('error')}")
            
            else:
                print(f"\n❌ {t('invalid_choice')}")
        
        elif choice == '3':
            resp = client.admin_report()
            if resp and resp.get('status') == 'success':
                print(f"\n📊 {t('admin_report_title')}")
                print("=" * 45)
                print(f"   👥 {t('total_users')} {resp.get('users_count', 0)}")
                print(f"   🏦 {t('total_accounts')} {resp.get('accounts_count', 0)}")
                print(f"   💰 {t('total_balance')} {resp.get('total_balance', 0):,.2f} {t('jod')}")
                
                stats = resp.get('stats', [])
                if stats:
                    print(f"\n📈 {t('transaction_stats')}")
                    for stat in stats:
                        if isinstance(stat, dict):
                            type_val = stat.get('type', '')
                            type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(type_val, type_val)
                            print(f"   {type_en}: {stat.get('count', 0)} {t('transactions')} | Total: {stat.get('total', 0):,.2f}")
                        else:
                            type_en = {'deposit': 'Deposit', 'withdraw': 'Withdraw', 'transfer': 'Transfer'}.get(stat[0] if len(stat) > 0 else '', stat[0] if len(stat) > 0 else '')
                            print(f"   {type_en}: {stat[1] if len(stat) > 1 else 0} {t('transactions')} | Total: {stat[2] if len(stat) > 2 else 0:,.2f}")
                
                active = resp.get('active_accounts', [])
                if active:
                    print(f"\n⭐ {t('most_active')}")
                    for acc in active[:5]:
                        if isinstance(acc, dict):
                            print(f"   {acc.get('account_number', '')}: {acc.get('transaction_count', 0)} {t('transactions')}")
                        else:
                            print(f"   {acc[0] if len(acc) > 0 else ''}: {acc[1] if len(acc) > 1 else 0} {t('transactions')}")
            else:
                print(f"\n❌ {resp.get('message') if resp else t('error')}")
        
        elif choice == '4':
            print(f"\n📊 {t('preparing_graph')}")
            resp = client.graph_report()
            if resp and resp.get('status') == 'success':
                show_graph_report(resp)
                send_notification('graph_report', 'Report generated successfully')
            else:
                print(f"\n❌ {t('graph_error')}: {resp.get('message') if resp else t('error')}")
        
        elif choice == '5':
            print(f"\n💾 {t('creating_excel')}")
            resp = client.backup()
            if resp and resp.get('status') == 'success':
                print(f"\n✅ {resp['message']}")
                send_notification('backup', resp['message'])
            else:
                print(f"\n❌ {resp.get('message') if resp else t('error')}")
        
        elif choice == '6':
            acc = input(f"{t('account_number')} ")
            try:
                limit = float(input(f"{t('daily_limit_text')} "))
                resp = client.set_daily_limit(acc, limit)
                if resp and resp.get('status') == 'success':
                    print(f"\n✅ {resp['message']}")
                    send_notification('set_limit', f"{acc}: {limit} JOD")
                else:
                    print(f"\n❌ {resp.get('message') if resp else t('error')}")
            except ValueError:
                print(f"\n❌ {t('invalid_amount')}")
        
        elif choice == '7':
            print(f"\n💱 {t('current_rates')}")
            resp = client.get_exchange_rates()
            if resp and resp.get('status') == 'success':
                rates = resp.get('rates', {})
                for curr, rate in rates.items():
                    print(f"   {curr}: 1 USD = {rate} {curr}")
                print(f"\n✏️ {t('enter_new_rates')}")
                new_rates = {}
                for curr in rates.keys():
                    val = input(f"{curr} {t('new_rate')} ")
                    if val.strip():
                        try:
                            new_rates[curr] = float(val)
                        except ValueError:
                            print(f"   Invalid rate, skipping {curr}")
                if new_rates:
                    resp2 = client.update_exchange_rates(new_rates)
                    if resp2 and resp2.get('status') == 'success':
                        print(f"\n✅ {resp2['message']}")
                        send_notification('update_rates', 'Exchange rates updated')
                    else:
                        print(f"\n❌ {resp2.get('message') if resp2 else t('error')}")
            else:
                print(f"\n❌ {resp.get('message') if resp else t('error')}")
        
        elif choice == '8':
            print(f"\n📎 {t('creating_excel')}")
            success, result = client.export_to_excel_file()
            if success:
                print(f"\n✅ {t('excel_created')}")
                print(f"    File: {result}")
                print(f"    Path: {os.path.abspath(result)}")
                send_notification('export_excel', result)
            else:
                print(f"\n❌ {t('excel_failed')}: {result}")
        
        elif choice == '9':
            print(f"\n 1. {t('multi_currency')}")
            print(f" 2. {t('convert')}")
            print("-" * 40)
            sub = input("Choose / اختر (1-2): ")
            
            if sub == '1':
                acc = input(f"{t('account_number')} ")
                resp = client.balance_multiple_currencies(acc)
                if resp and resp.get('status') == 'success':
                    print(f"\n{t('balance_jod')} {resp['balance_jod']:,.2f} JD")
                    print(f"\n{t('other_currencies')}")
                    print("-" * 35)
                    for curr, amt in resp.get('balances', {}).items():
                        print(f"   {curr:<6}: {amt:>15,.2f}")
                    print("-" * 35)
                else:
                    print(f"\n❌ {resp.get('message') if resp else t('error')}")
            
            elif sub == '2':
                try:
                    amt = float(input(f"{t('amount')} "))
                    print(f"\n{t('available_currencies')}")
                    frm = input(f"{t('from_currency')} ").upper()
                    to = input(f"{t('to_currency')} ").upper()
                    resp = client.convert_currency(amt, frm, to)
                    if resp and resp.get('status') == 'success':
                        print(f"\n✅ {t('convert_result')}")
                        print(f"   {amt:,.2f} {frm} = {resp['converted_amount']:,.2f} {to}")
                        print(f"   Exchange Rate: {resp['rate']}")
                    else:
                        print(f"\n❌ {resp.get('message') if resp else t('error')}")
                except ValueError:
                    print(f"\n❌ {t('invalid_amount')}")
            
            else:
                print(f"\n❌ {t('invalid_choice')}")
        
        elif choice == '10':
            old = getpass.getpass(f"{t('old_password')} ")
            new = getpass.getpass(f"{t('new_password_text')} ")
            conf = getpass.getpass(f"{t('confirm_new')} ")
            if new != conf:
                print(f"\n❌ {t('password_mismatch')}")
            else:
                resp = client.change_password(client.username, old, new)
                if resp and resp.get('status') == 'success':
                    print(f"\n✅ {resp['message']}")
                    send_notification('change_pass', 'Password changed successfully')
                else:
                    print(f"\n❌ {resp.get('message') if resp else t('error')}")
        
        elif choice == '11':
            print(f"\n{t('logout')}")
            client.logout()
            print(f"✅ {t('logged_out')}")
            send_notification('logout', 'Logged out successfully')
            input(f"\n{t('press_enter')}")
            return
        
        else:
            print(f"\n❌ {t('invalid_choice')}")
        
        input(f"\n{t('press_enter')}")

# ------------------------------------------------------------
# Main Function
# ------------------------------------------------------------
def main():
    # Show language selection first
    select_language()
    
    while True:
        clear_screen()
        print("=" * 55)
        print(f"   {t('app_title')}")
        print("=" * 55)
        print()
        print(f"{t('demo_accounts')}")
        print(f"   admin   | admin123 | {t('admin_role')}")
        print(f"   teller  | teller123 | {t('teller_role')}")
        print(f"   ahmed   | 123456 | {t('customer_role')} ({t('account_note')})")
        print()
        print(f"{t('tip_title')}")
        print(f"   {t('tip1')}")
        print(f"   {t('tip2')}")
        print(f"   {t('tip3')}")
        print()
        
        client = BankClient()
        
        if not client.connect():
            print(f"❌ {t('connecting_error')}")
            input(f"\n{t('press_enter')}")
            continue
        
        while not client.authenticated:
            print("\n" + "=" * 40)
            username = input(f"{t('username')} ")
            password = getpass.getpass(f"{t('password')} ")
            
            success, message = client.login(username, password)
            if success:
                print(f"\n✅ {message}")
                print(f"    {t('welcome')} {username} - {t('role')}: {client.role}")
                send_notification('login', f"Welcome {username}")
                input(f"\n{t('press_enter')}")
                
                if client.role == 'customer':
                    account_number = input(f"\n{t('enter_account')} ")
                    customer_menu(client, account_number)
                elif client.role == 'teller':
                    teller_menu(client)
                elif client.role == 'admin':
                    admin_menu(client)
                else:
                    print(f"\n❌ {t('invalid_choice')}")
                    client.logout()
                
                try:
                    client.socket.close()
                except:
                    pass
                break
            else:
                print(f"\n❌ {message}")
                input(f"\n{t('press_enter')}")
        
        continue

if __name__ == '__main__':
    main()
